from django.shortcuts import render
from django.core.mail import send_mail
from django.http import HttpResponse,JsonResponse
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework.status import HTTP_201_CREATED,HTTP_302_FOUND
from rest_framework import status
from django.conf import settings
from . models import LoginData,BatchData,StudentData,DateData,SendOTP
from . serializer import LoginDataSerializer,BatchDataSerializer,StudentDataSerializer,DateDataSerializer,SendOTPSerializer
from openpyxl import load_workbook

# Create your views here.

@api_view(['GET','POST','PUT','DELETE'])
def Login_Data(request):
    if request.method == 'GET':
        Item_Data = LoginData.objects.all()
        serializer = LoginDataSerializer(Item_Data, many=True)
        return Response(serializer.data)

    if request.method == 'POST':
        Data = LoginDataSerializer(data = request.data)
        if Data.is_valid() == True:
            Data.save()
        return Response(status= HTTP_201_CREATED)
    
    elif request.method == 'PUT':
        try:
            item = LoginData.objects.get(pk=request.data.get('id'))
        except LoginData.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)

        serializer = LoginDataSerializer(item, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        try:
            item = LoginData.objects.get(pk=request.data.get('id'))
        except LoginData.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)

        item.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    

@api_view(['GET', 'POST', 'PUT', 'DELETE'])
def Batch_Data(request):
    if request.method == 'GET':
        items = BatchData.objects.all()
        serializer = BatchDataSerializer(items, many=True)
        return Response(serializer.data)

    elif request.method == 'POST':
        serializer = BatchDataSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'PUT':
        try:
            item = BatchData.objects.get(pk=request.data.get('id'))
        except BatchData.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)

        serializer = BatchDataSerializer(item, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        try:
            item = BatchData.objects.get(pk=request.data.get('id'))
        except BatchData.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)

        item.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    

@api_view(['GET','POST','PUT','DELETE'])
def Student_Data(request):
    if request.method == 'GET':
        Item_Data = StudentData.objects.all()
        serializer = StudentDataSerializer(Item_Data, many=True)
        return Response(serializer.data)

    if request.method == 'POST':
        Data = StudentDataSerializer(data = request.data)
        print(request.data)
        if Data.is_valid() == True:
            Data.save()
            return Response(status=HTTP_201_CREATED)
        return Response(status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'PUT':
        try:
            item = StudentData.objects.get(pk=request.data.get('id'))
        except StudentData.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)

        serializer = StudentDataSerializer(item, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data,status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        try:
            item = StudentData.objects.get(pk=request.data.get('id'))
        except StudentData.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)

        item.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    
@api_view(['GET','POST'])
def StudentFiles_Data(request):
    if request.method == 'POST':
        try:
            excel_file = request.FILES['excel_file']
            wb = load_workbook(excel_file, data_only=True)

            sheet = wb.active
            data_list = []

            for row in sheet.iter_rows(min_row=2):
                data = {}

                for cell in row:
                    column_name = sheet.cell(row=1, column=cell.column).value.lower()

                    if column_name == 'batchname':
                        data['BatchName'] = cell.value
                    elif column_name == 'joiningdate':
                        date = str(cell.value)[0:11]
                        fDate = f"{date[8:10]}-{date[5:7]}-{date[:4]}"
                        data['JoiningDate'] = fDate
                    elif column_name == 'name':
                        data['Name'] = cell.value
                    elif column_name == 'email':
                        data['Email'] = cell.value
                    elif column_name == 'phone':
                        data['Phone'] = cell.value

                if data.get('BatchName') and data.get('JoiningDate') is not None:
                    if data.get('Name') and data.get('Email') is not None and data.get('Phone') is not None:
                        data_list.append(data)
                
            serializer = StudentDataSerializer(data=data_list, many=True)
            serializer.is_valid(raise_exception=True)

            serializer.save()

            return Response({'message': 'Data uploaded successfully!'}, status=201)
        except Exception as e:
            print(f"Error uploading data: {e}")
            return Response({'error': str(e)}, status=400)

@api_view(['GET','POST','PUT','DELETE'])
def Date_Data(request):
    if request.method == 'GET':
        Item_Data = DateData.objects.all()
        serializer = DateDataSerializer(Item_Data, many=True)
        return Response(serializer.data)

    if request.method == 'POST':
        Data = DateDataSerializer(data = request.data)
        if Data.is_valid() == True:
            Data.save()
        return Response(status=HTTP_201_CREATED)

    elif request.method == 'PUT':
        try:
            item = DateData.objects.get(pk=request.data.get('id'))
        except DateData.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)

        serializer = DateDataSerializer(item, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        try:
            item = DateData.objects.get(pk=request.data.get('id'))
        except DateData.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)

        item.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    
    
def contains_alphanumeric(s):
    return any(char.isalnum() for char in s)

@api_view(['GET', 'POST'])
def Send_OTP(request):
    if request.method == 'GET':
        Item_Data = SendOTP.objects.all()
        serializer = SendOTPSerializer(Item_Data, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    if request.method == 'POST':
        serializer = SendOTPSerializer(data=request.data)
        if serializer.is_valid():
            email = serializer.validated_data.get('Email')
            mailtype = serializer.validated_data.get('OTP').split(' ')[-1]
            if (mailtype == 'Auth_Alert' or mailtype == 'Std_Login_Alert' or mailtype == 'Std_Details_Update_Alert'):
                replaceotp = serializer.validated_data.get('OTP').split(' ')[1].replace('}',"")
                otp = f'{serializer.validated_data.get('OTP').split(' ')[0]} {replaceotp}'
            else:    
                otp = serializer.validated_data.get('OTP').split(' ')[0]
                
            if mailtype == 'Username':
                title = 'Forgot Username'
                message = f'''Dear User,
                
Your username is {otp}.
Please keep this information secure and do not share it with anyone. If you did not request this information, please disregard this message and ensure your account security is intact.

If you have any concerns or questions, feel free to contact our support team.

Thank you,
VCUBE Support Team.
'''
            elif mailtype == 'OTP':
                title = 'Forgot Password'
                name = 'reset your password'
                name2 = 'password reset process'
                
            elif mailtype == 'User_OTP':
                title = 'New User'
                name = 'add a new user'
                name2 = 'new user add process'
            
            elif mailtype == 'User_Auth':
                title = 'Change Authentication'
                name = 'change authentication'
                name2 = 'change authentication process'
                
            elif mailtype == 'User_Details_Change':
                title = 'Change User Details'
                name = 'change user details'
                name2 = 'change user details'
                
            elif mailtype == 'Std_Login_OTP':
                title = 'User Login'
                name = 'login'
                name2 = 'login process'
                
            elif mailtype == 'Auth_Alert':
                title = 'Student Authentication Alert !'
                name = f'We have detected a change in your login authentication status at {otp}'   

            elif mailtype == 'Std_Login_Alert':
                title = 'Student Login Alert !'
                name = f'We have detected a login attempt to your account at {otp}'

            elif mailtype == 'Std_Details_Update_Alert':
                title = 'Student Details Update Alert !'
                name = f'We have detected an update to your student details on {otp}'
            
            if mailtype == 'OTP' or mailtype == 'User_OTP' or mailtype == 'User_Auth' or mailtype == 'User_Details_Change' or mailtype == 'Std_Login_OTP':
                message = f'''Dear User,
                
Please use the following OTP code : {otp} to {name}.
Ensure to input the code accurately to proceed with the {name2}. If you did not initiate this request, please disregard and ensure your account security is intact.

If you have any concerns or questions, feel free to contact our support team.

Thank you,
VCUBE Support Team.
'''     
            
            elif mailtype == 'Auth_Alert' or mailtype == 'Std_Login_Alert' or mailtype == 'Std_Details_Update_Alert':
                message = f'''Dear Student,

{name}. If this was not you, it is crucial that you contact our support team immediately to secure your account. Unauthorized access can compromise your personal information and academic records. 
Please ensure your account is safe by reaching out to support at https://www.vcubesoftsolutions.com/ or you can contact directly at institute VCube Software Solutions, Hyderabad. 

Thank you for your prompt attention to this matter.
VCUBE Support Team.
'''  

            send_mail(
                title,
                message,
                settings.EMAIL_HOST_USER,
                [email],
                fail_silently=False
            )
            return Response(status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)    